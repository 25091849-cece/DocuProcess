from flask import Flask, render_template_string, request, redirect, url_for, jsonify, send_file
import boto3
from botocore.exceptions import ClientError
from werkzeug.utils import secure_filename
import uuid
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

app = Flask(__name__)
dynamodb = boto3.resource('dynamodb', region_name='us-east-1')
s3_client = boto3.client('s3', region_name='us-east-1')
table = dynamodb.Table('JusticeArchDocuments')

BUCKET_NAME = 'justicearch-inbox-group7'   # ← replace with your actual bucket name

def get_presigned_url(s3_key):
    """Generate a temporary URL so admin can open the PDF in browser."""
    try:
        url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': BUCKET_NAME, 'Key': s3_key},
            ExpiresIn=300   # URL valid for 5 minutes
        )
        return url
    except ClientError:
        return None

TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
  <title>JusticeArch Review Portal</title>
  <style>
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: Arial, sans-serif; background: #f0f4f8; color: #222; }

    /* ── Top bar ── */
    .topbar {
      background: #1F4E79; color: white;
      padding: 16px 32px; display: flex;
      align-items: center; justify-content: space-between;
    }
    .topbar h1 { font-size: 20px; }
    .topbar span { font-size: 13px; opacity: 0.75; }

    /* ── Stats row ── */
    .stats {
      display: flex; gap: 16px;
      padding: 20px 32px; background: #e8f0f7;
      border-bottom: 1px solid #ccd9e8;
    }
    .stat-box {
      background: white; border-radius: 8px;
      padding: 14px 24px; text-align: center;
      box-shadow: 0 1px 4px rgba(0,0,0,0.08); min-width: 130px;
    }
    .stat-box .num { font-size: 28px; font-weight: bold; }
    .stat-box .lbl { font-size: 12px; color: #666; margin-top: 4px; }
    .num.green  { color: #1a7f4b; }
    .num.red    { color: #c0392b; }
    .num.orange { color: #e67e22; }
    .num.blue   { color: #1F4E79; }

    /* ── Main content ── */
    .content { padding: 24px 32px; }

    /* ── Search bar ── */
    .searchbar { display: flex; gap: 10px; margin-bottom: 20px; }
    .searchbar input {
      flex: 1; padding: 10px 14px; border: 1px solid #b0c4d8;
      border-radius: 6px; font-size: 14px;
    }
    .searchbar button {
      padding: 10px 22px; background: #1F4E79; color: white;
      border: none; border-radius: 6px; cursor: pointer; font-size: 14px;
    }
    .searchbar a {
      padding: 10px 18px; background: #6c757d; color: white;
      border-radius: 6px; text-decoration: none; font-size: 14px;
    }
    .searchbar .btn-refresh {
      padding: 10px 18px; background: #17a2b8; color: white;
      border: none; border-radius: 6px; cursor: pointer; font-size: 14px;
      transition: background 0.3s;
    }
    .searchbar .btn-refresh:hover {
      background: #138496;
    }

    /* ── Sort Controls ── */
    .sort-controls {
      display: flex; gap: 12px; margin-bottom: 20px;
      align-items: center; padding: 12px 16px;
      background: #f0f4f8; border-radius: 8px;
      border: 1px solid #b0c4d8;
    }
    .sort-btn {
      padding: 8px 16px; background: white; color: #1F4E79;
      border: 2px solid #b0c4d8; border-radius: 6px;
      text-decoration: none; font-size: 13px; font-weight: 600;
      cursor: pointer; transition: all 0.3s;
    }
    .sort-btn:hover {
      background: #e8f0f7; border-color: #1F4E79;
    }
    .sort-btn-active {
      background: #1F4E79; color: white; border-color: #1F4E79;
    }
    .sort-btn-active:hover {
      background: #153556;
    }

    /* ── Table ── */
    .table-wrap { overflow-x: auto; background: white;
                  border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
    table { width: 100%; border-collapse: collapse; min-width: 900px; }
    th {
      background: #1F4E79; color: white;
      padding: 12px 10px; text-align: left; font-size: 13px;
    }
    td { padding: 10px; border-bottom: 1px solid #e0eaf4; font-size: 13px; vertical-align: middle; }
    tr:last-child td { border-bottom: none; }
    tr:hover td { background: #f5f9ff; }

    /* ── Badges ── */
    .badge {
      display: inline-block; padding: 3px 10px; border-radius: 12px;
      font-size: 11px; font-weight: bold; text-transform: uppercase;
    }
    .badge-approved { background: #d4edda; color: #1a7f4b; }
    .badge-pending  { background: #fff3cd; color: #856404; }
    .badge-rejected { background: #f8d7da; color: #842029; }

    .conf-high   { color: #1a7f4b; font-weight: bold; }
    .conf-mid    { color: #e67e22; font-weight: bold; }
    .conf-low    { color: #c0392b; font-weight: bold; }

    /* High-priority row highlight (confidence < 60%) */
    tr.priority-high td { background: #fff0f0 !important; }
    tr.priority-high:hover td { background: #ffe0e0 !important; }
    tr.priority-high { border-left: 4px solid #dc3545; }

    /* ── Action buttons ── */
    .btn {
      padding: 6px 12px; border: none; border-radius: 5px;
      cursor: pointer; font-size: 12px; font-weight: bold;
      text-decoration: none; display: inline-block; margin: 2px;
    }
    .btn-view     { background: #0d6efd; color: white; }
    .btn-approve  { background: #198754; color: white; }
    .btn-reject   { background: #dc3545; color: white; }
    .btn-disabled { background: #ccc; color: #888; cursor: not-allowed; }

    /* ── Modal overlay ── */
    .modal-overlay {
      display: none; position: fixed; inset: 0;
      background: rgba(0,0,0,0.65); z-index: 1000;
      align-items: center; justify-content: center;
    }
    .modal-overlay.active { display: flex; }
    .modal {
      background: white; border-radius: 12px; width: 90vw; max-width: 980px;
      max-height: 92vh; display: flex; flex-direction: column;
      box-shadow: 0 8px 40px rgba(0,0,0,0.35); overflow: hidden;
    }
    .modal-header {
      background: #1F4E79; color: white;
      padding: 14px 20px; display: flex;
      justify-content: space-between; align-items: center;
    }
    .modal-header h3 { font-size: 16px; }
    .modal-close {
      background: none; border: none; color: white;
      font-size: 22px; cursor: pointer; line-height: 1;
    }
    .modal-body { flex: 1; overflow: hidden; display: flex; flex-direction: column; }
    .modal-body iframe { flex: 1; width: 100%; border: none; min-height: 500px; }

    /* ── PDF info panel inside modal ── */
    .pdf-meta {
      padding: 12px 18px; background: #f8f9fa;
      border-bottom: 1px solid #dee2e6;
      display: flex; flex-wrap: wrap; gap: 18px; font-size: 13px;
    }
    .pdf-meta span b { color: #1F4E79; }

    /* ── Reject form inside modal ── */
    .reject-section {
      padding: 14px 18px; background: #fff8f8;
      border-top: 2px solid #f8d7da;
    }
    .reject-section h4 { color: #842029; margin-bottom: 10px; font-size: 14px; }
    .reject-section textarea {
      width: 100%; padding: 8px; border: 1px solid #f5c2c7;
      border-radius: 6px; font-size: 13px; resize: vertical; min-height: 70px;
    }
    .reject-actions { display: flex; gap: 10px; margin-top: 10px; }

    /* ── Approve section inside modal ── */
    .approve-section {
      padding: 14px 18px; background: #f0fff4;
      border-top: 2px solid #d4edda; display: flex;
      align-items: center; justify-content: space-between;
    }
    .approve-section p { font-size: 13px; color: #1a7f4b; }

    /* ── Alert banner ── */
    .alert {
      padding: 12px 18px; border-radius: 8px; margin-bottom: 18px; font-size: 14px;
    }
    .alert-info { background: #cfe2ff; border: 1px solid #9ec5fe; color: #084298; }

    /* ── Checkbox styling ── */
    input[type="checkbox"] {
      cursor: pointer; width: 18px; height: 18px; margin: 0;
    }
    th input[type="checkbox"] { margin: 2px; }

    /* ── Selection toolbar ── */
    .selection-toolbar {
      display: none; padding: 14px 32px; background: #e3f2fd;
      border-bottom: 2px solid #1F4E79; margin-bottom: 16px;
      align-items: center; justify-content: space-between;
    }
    .selection-toolbar.active { display: flex; }
    .selection-toolbar .info { font-size: 14px; color: #1F4E79; font-weight: bold; }
    .selection-toolbar .actions { display: flex; gap: 10px; }
    .btn-delete-selected {
      background: #dc3545; color: white; padding: 8px 18px;
      border: none; border-radius: 6px; cursor: pointer;
      font-weight: bold; font-size: 13px;
    }
    .btn-delete-selected:hover { background: #c82333; }
    .btn-cancel-selection {
      background: #6c757d; color: white; padding: 8px 16px;
      border: none; border-radius: 6px; cursor: pointer; font-size: 13px;
    }
    .btn-cancel-selection:hover { background: #5a6268; }

    /* ── Confirmation Modal ── */
    .confirm-modal-overlay {
      display: none; position: fixed; inset: 0;
      background: rgba(0,0,0,0.7); z-index: 2000;
      align-items: center; justify-content: center;
    }
    .confirm-modal-overlay.active { display: flex; }
    .confirm-modal {
      background: white; border-radius: 12px; padding: 28px;
      width: 90%; max-width: 450px; text-align: center;
      box-shadow: 0 10px 50px rgba(0,0,0,0.3);
    }
    .confirm-modal h2 {
      color: #dc3545; font-size: 20px; margin-bottom: 12px;
    }
    .confirm-modal p {
      color: #555; font-size: 14px; line-height: 1.6; margin-bottom: 20px;
    }
    .confirm-modal .count {
      font-weight: bold; color: #1F4E79; font-size: 16px;
    }
    .confirm-actions {
      display: flex; gap: 12px; justify-content: center;
    }
    .btn-confirm-delete {
      background: #dc3545; color: white; padding: 10px 24px;
      border: none; border-radius: 6px; cursor: pointer;
      font-weight: bold; font-size: 14px;
    }
    .btn-confirm-delete:hover { background: #c82333; }
    .btn-confirm-cancel {
      background: #e0e0e0; color: #333; padding: 10px 24px;
      border: none; border-radius: 6px; cursor: pointer; font-size: 14px;
    }
    .btn-confirm-cancel:hover { background: #ccc; }

    tr.selected { background: #fff3cd !important; }
    tr.selected td { background: #fff3cd; }

    /* ── Tab Navigation ── */
    .tab-nav {
      display: flex; gap: 0; border-bottom: 2px solid #1F4E79;
      margin-bottom: 20px; background: white;
    }
    .tab-button {
      padding: 12px 24px; background: #f0f4f8; border: none;
      border-bottom: 3px solid transparent; cursor: pointer; font-size: 14px;
      font-weight: 600; color: #555; transition: all 0.3s;
    }
    .tab-button:hover { background: #e0eaf4; }
    .tab-button.active {
      background: white; color: #1F4E79; border-bottom-color: #1F4E79;
    }
    .tab-content { display: none; }
    .tab-content.active { display: block; }

    /* ── Upload Area ── */
    .upload-container {
      background: white; border-radius: 10px; padding: 32px;
      box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    }
    .upload-instructions {
      background: #e8f0f7; border-left: 4px solid #1F4E79;
      padding: 16px; border-radius: 6px; margin-bottom: 24px;
    }
    .upload-instructions h3 {
      color: #1F4E79; font-size: 15px; margin-bottom: 8px;
    }
    .upload-instructions ul {
      list-style: none; padding-left: 0; font-size: 13px; color: #555;
    }
    .upload-instructions li {
      padding: 4px 0; margin-left: 20px;
    }
    .upload-instructions li:before {
      content: "✓ "; color: #1a7f4b; font-weight: bold; margin-left: -20px;
      margin-right: 8px;
    }

    .file-input-wrapper {
      position: relative; display: inline-block;
    }
    .file-input-label {
      display: inline-block; padding: 14px 28px; background: #1F4E79;
      color: white; border-radius: 6px; cursor: pointer; font-weight: bold;
      transition: background 0.3s;
    }
    .file-input-label:hover { background: #153556; }
    .file-input-wrapper input[type="file"] {
      display: none;
    }

    .file-list {
      margin-top: 20px;
    }
    .file-item {
      display: flex; align-items: center; justify-content: space-between;
      padding: 12px 16px; background: #f8f9fa; border: 1px solid #dee2e6;
      border-radius: 6px; margin-bottom: 10px;
    }
    .file-item-info {
      flex: 1;
    }
    .file-item-name {
      font-weight: 600; color: #222; font-size: 13px;
    }
    .file-item-size {
      font-size: 12px; color: #666; margin-top: 4px;
    }
    .file-item-remove {
      background: #dc3545; color: white; border: none;
      padding: 6px 12px; border-radius: 4px; cursor: pointer; font-size: 12px;
    }
    .file-item-remove:hover { background: #c82333; }

    .upload-progress {
      display: none; margin-top: 20px;
    }
    .upload-progress.active { display: block; }
    .progress-bar {
      width: 100%; height: 24px; background: #e0e0e0;
      border-radius: 6px; overflow: hidden; margin-bottom: 8px;
    }
    .progress-fill {
      height: 100%; background: linear-gradient(90deg, #1F4E79, #2a6da6);
      width: 0%; transition: width 0.3s;
    }
    .progress-text {
      font-size: 12px; color: #666; text-align: center;
    }

    .upload-success {
      display: none; padding: 16px; background: #d4edda;
      border: 1px solid #c3e6cb; border-radius: 6px; color: #1a7f4b;
      margin-top: 20px; text-align: center;
    }
    .upload-success.active { display: block; }

    .upload-error {
      display: none; padding: 16px; background: #f8d7da;
      border: 1px solid #f5c2c7; border-radius: 6px; color: #842029;
      margin-top: 20px;
    }
    .upload-error.active { display: block; }

    .btn-upload {
      padding: 12px 28px; background: #198754; color: white;
      border: none; border-radius: 6px; cursor: pointer;
      font-weight: bold; font-size: 14px; margin-top: 20px;
    }
    .btn-upload:hover { background: #157347; }
    .btn-upload:disabled {
      background: #6c757d; cursor: not-allowed;
    }

    .file-count-info {
      font-size: 13px; color: #666; margin-top: 16px;
      padding: 12px; background: #fff3cd; border-radius: 6px;
    }
  </style>
</head>
<body>

<!-- Top bar -->
<div class="topbar">
  <h1>☁ JusticeArch Document Review Portal</h1>
  <span>Powered by AWS Textract + DynamoDB</span>
</div>

<!-- Stats -->
<div class="stats">
  <div class="stat-box">
    <div class="num blue">{{ total }}</div>
    <div class="lbl">Total Documents</div>
  </div>
  <div class="stat-box">
    <div class="num green">{{ approved }}</div>
    <div class="lbl">Auto-Approved</div>
  </div>
  <div class="stat-box">
    <div class="num orange">{{ pending }}</div>
    <div class="lbl">Pending Review</div>
  </div>
  <div class="stat-box">
    <div class="num red">{{ rejected }}</div>
    <div class="lbl">Rejected</div>
  </div>
</div>

<!-- Main content -->
<div class="content">

  <!-- Tab Navigation -->
  <div class="tab-nav">
    <button class="tab-button active" onclick="switchTab('review-tab')">
      📋 Document Review
    </button>
    <button class="tab-button" onclick="switchTab('upload-tab')">
      📤 Upload Documents
    </button>
  </div>

  <!-- ==================== REVIEW TAB ==================== -->
  <div id="review-tab" class="tab-content active">

    <!-- Search -->
    <form class="searchbar" method="GET" action="/search">
      <button type="button" class="btn-refresh" onclick="window.location.href='/';" title="Refresh data">
        🔄 Refresh
      </button>
      <input name="q" placeholder="Search by vendor, date, amount, status..."
             value="{{ query or '' }}">
      <button type="submit">🔍 Search</button>
      <a href="/">Clear</a>
      <a href="/download-excel" style="background: #28a745;" title="Download all documents as Excel">📊 Download Excel</a>
    </form>

    <!-- Sort Controls -->
    <div class="sort-controls">
      <label style="font-weight:600; margin-right:12px; color:#1F4E79;">Sort by Date:</label>
      <a href="?q={{ query or '' }}&sort=desc" 
         class="sort-btn {% if sort_order == 'desc' %}sort-btn-active{% endif %}"
         title="Sort from newest to oldest">
        📅 Newest First (↓ Descending)
      </a>
      <a href="?q={{ query or '' }}&sort=asc" 
         class="sort-btn {% if sort_order == 'asc' %}sort-btn-active{% endif %}"
         title="Sort from oldest to newest">
        📅 Oldest First (↑ Ascending)
      </a>
    </div>

    {% if query %}
    <div class="alert alert-info">
      Showing results for: <b>"{{ query }}"</b> — {{ docs|length }} record(s) found.
    </div>
    {% endif %}

    <!-- Selection toolbar (hidden by default) -->
    <div class="selection-toolbar" id="selectionToolbar">
      <div class="info">
        <span id="selectedCount">0</span> document(s) selected
      </div>
      <div class="actions">
        <button type="button" class="btn-delete-selected" onclick="confirmDelete()">
          🗑️ Delete Selected
        </button>
        <button type="button" class="btn-cancel-selection" onclick="clearSelection()">
          Cancel
        </button>
      </div>
    </div>

    <!-- Table -->
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th style="width:50px;text-align:center">
              <input type="checkbox" id="selectAllCheckbox" onchange="toggleSelectAll(this)">
            </th>
            <th>Doc ID</th>
            <th>File Name</th>
            <th>Vendor</th>
            <th>Date</th>
            <th>Amount</th>
            <th>Confidence</th>
            <th>Status</th>
            <th>Uploaded</th>
            <th style="text-align:center">Actions</th>
          </tr>
        </thead>
        <tbody id="tableBody">
      {% for doc in docs %}
        <tr class="doc-row {% if doc.reviewPriority == 'HIGH' %}priority-high{% endif %}" data-doc-id="{{ doc.documentId }}">
          <td style="text-align:center">
            <input type="checkbox" class="row-checkbox" onchange="updateSelection()">
          </td>
          <td><code style="font-size:11px">{{ doc.documentId[:8] }}...</code></td>
          <td style="max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"
              title="{{ doc.s3Key }}">{{ doc.s3Key }}</td>
          <td>{{ doc.vendor }}</td>
          <td>{{ doc.date }}</td>
          <td><b>{{ doc.amount }}</b></td>
          <td>
            {% set conf = doc.confidence | float %}
            {% if conf >= 80 %}
              <span class="conf-high">{{ doc.confidence }}% ✅</span>
            {% elif conf >= 60 %}
              <span class="conf-mid">{{ doc.confidence }}% ⚠️</span>
            {% else %}
              <span class="conf-low">{{ doc.confidence }}% 🔴</span>
            {% endif %}
          </td>
          <td>
            {% if doc.status == 'APPROVED' %}
              <span class="badge badge-approved">✓ Approved</span>
            {% elif doc.status == 'REJECTED' %}
              <span class="badge badge-rejected">✗ Rejected</span>
              {% if doc.rejectReason %}
                <div style="font-size:11px;color:#842029;margin-top:4px">
                  Reason: {{ doc.rejectReason }}
                </div>
              {% endif %}
            {% else %}
              <span class="badge badge-pending">⏳ Pending</span>
              {% if doc.reviewPriority == 'HIGH' %}
                <div style="font-size:11px;color:#dc3545;font-weight:bold;margin-top:4px">
                  🔴 优先审核
                </div>
              {% endif %}
            {% endif %}
          </td>
          <td style="font-size:11px;color:#666">
            {{ doc.uploadedAt[:10] if doc.uploadedAt else '-' }}
          </td>
          <td style="text-align:center;white-space:nowrap">
            {% if doc.flaggedForReview and doc.status == 'PENDING' %}
              <!-- Low confidence: show View + Approve + Reject -->
              <button class="btn btn-view"
                onclick="openModal(
                  '{{ doc.documentId }}',
                  '{{ doc.s3Key }}',
                  '{{ doc.vendor }}',
                  '{{ doc.date }}',
                  '{{ doc.amount }}',
                  '{{ doc.confidence }}',
                  '{{ doc.pdfUrl }}'
                )">📄 View PDF</button>
            {% elif doc.status == 'PENDING' %}
              <!-- High conf but still pending -->
              <form method="POST" action="/approve" style="display:inline">
                <input type="hidden" name="id" value="{{ doc.documentId }}">
                <button class="btn btn-approve">✓ Approve</button>
              </form>
            {% else %}
              <span class="btn btn-disabled">{{ doc.status }}</span>
            {% endif %}
          </td>
        </tr>
      {% endfor %}
      {% if not docs %}
        <tr>
          <td colspan="10" style="text-align:center;padding:30px;color:#888">
            No documents found.
          </td>
        </tr>
      {% endif %}
      </tbody>
    </table>
  </div>

  </div><!-- End Review Tab -->

  <!-- ==================== UPLOAD TAB ==================== -->
  <div id="upload-tab" class="tab-content">

    <div class="upload-container">

      <div class="upload-instructions">
        <h3>📤 Upload PDF Documents to S3</h3>
        <ul>
          <li>Maximum 3 PDF files per upload</li>
          <li>Maximum file size: 10 MB per file</li>
          <li>Files will be securely uploaded to AWS S3</li>
          <li>Processing will begin automatically after upload</li>
        </ul>
      </div>

      <div class="file-input-wrapper">
        <label class="file-input-label" for="pdfFileInput">
          📁 Select PDF Files
        </label>
        <input type="file" id="pdfFileInput" multiple accept=".pdf" 
               onchange="handleFileSelect(event)">
      </div>

      <div class="file-list" id="fileList"></div>

      <div class="file-count-info" id="fileCountInfo" style="display:none;">
        <span id="selectedFileCount">0</span> file(s) selected
      </div>

      <button class="btn-upload" id="uploadBtn" onclick="uploadFiles()" disabled>
        🚀 Upload Files
      </button>

      <div class="upload-progress" id="uploadProgress">
        <div class="progress-bar">
          <div class="progress-fill" id="progressFill"></div>
        </div>
        <div class="progress-text" id="progressText">0%</div>
      </div>

      <div class="upload-success" id="uploadSuccess">
        ✅ <span id="successMessage">Files uploaded successfully!</span>
      </div>

      <div class="upload-error" id="uploadError">
        ❌ <span id="errorMessage">Upload failed</span>
      </div>

    </div>

  </div><!-- End Upload Tab -->

</div><!-- End content -->

<!-- ── PDF Review Modal ─────────────────────────────────────────────────── -->
<div class="modal-overlay" id="pdfModal">
  <div class="modal">

    <div class="modal-header">
      <h3>📄 PDF Review — <span id="modalTitle"></span></h3>
      <button class="modal-close" onclick="closeModal()">✕</button>
    </div>

    <div class="modal-body">

      <!-- Metadata strip -->
      <div class="pdf-meta">
        <span><b>Vendor:</b> <span id="mVendor"></span></span>
        <span><b>Date:</b> <span id="mDate"></span></span>
        <span><b>Amount:</b> <span id="mAmount"></span></span>
        <span><b>Confidence:</b> <span id="mConf" style="color:#c0392b;font-weight:bold"></span>%</span>
        <span style="color:#856404">⚠️ Low confidence — please verify the PDF before approving.</span>
      </div>

      <!-- PDF viewer -->
      <iframe id="pdfFrame" src="" title="Invoice PDF"></iframe>

      <!-- Approve section -->
      <div class="approve-section">
        <p>✅ Document looks correct? Approve it below.</p>
        <form method="POST" action="/approve" id="approveForm">
          <input type="hidden" name="id" id="approveId">
          <button class="btn btn-approve" style="padding:10px 24px;font-size:14px">
            ✓ Approve Document
          </button>
        </form>
      </div>

      <!-- Reject section -->
      <div class="reject-section">
        <h4>✗ Reject this document</h4>
        <form method="POST" action="/reject" id="rejectForm">
          <input type="hidden" name="id" id="rejectId">
          <textarea name="reason"
            placeholder="Enter reason for rejection (e.g. Wrong vendor name, Amount mismatch, Duplicate invoice, Unreadable document...)"></textarea>
          <div class="reject-actions">
            <button class="btn btn-reject" type="submit"
                    style="padding:8px 20px;font-size:13px">
              ✗ Reject Document
            </button>
            <span style="font-size:12px;color:#888;align-self:center">
              A reason is recommended before rejecting.
            </span>
          </div>
        </form>
      </div>

    </div>
  </div>
</div>

<!-- ── Delete Confirmation Modal ───────────────────────────────────────── -->
<div class="confirm-modal-overlay" id="deleteConfirmModal">
  <div class="confirm-modal">
    <h2>⚠️ Delete Documents</h2>
    <p>
      Are you sure you want to permanently delete <span class="count" id="deleteCount">0</span> 
      document(s) from DynamoDB? This action cannot be undone.
    </p>
    <div class="confirm-actions">
      <button type="button" class="btn-confirm-delete" onclick="executeDelete()">
        Yes, Delete
      </button>
      <button type="button" class="btn-confirm-cancel" onclick="cancelDelete()">
        Cancel
      </button>
    </div>
  </div>
</div>

<script>
// ── Selection and Deletion Logic ──
function getSelectedDocIds() {
  const checkboxes = document.querySelectorAll('.row-checkbox:checked');
  return Array.from(checkboxes).map(cb => cb.closest('tr').dataset.docId);
}

function updateSelection() {
  const selectedIds = getSelectedDocIds();
  const toolbar = document.getElementById('selectionToolbar');
  const countSpan = document.getElementById('selectedCount');
  const allCheckbox = document.getElementById('selectAllCheckbox');
  const totalCheckboxes = document.querySelectorAll('.row-checkbox').length;
  const checkedCheckboxes = document.querySelectorAll('.row-checkbox:checked').length;

  countSpan.textContent = selectedIds.length;

  // Show/hide toolbar
  if (selectedIds.length > 0) {
    toolbar.classList.add('active');
  } else {
    toolbar.classList.remove('active');
  }

  // Update select-all checkbox state
  if (totalCheckboxes > 0) {
    allCheckbox.indeterminate = checkedCheckboxes > 0 && checkedCheckboxes < totalCheckboxes;
    allCheckbox.checked = checkedCheckboxes === totalCheckboxes && totalCheckboxes > 0;
  }

  // Highlight selected rows
  document.querySelectorAll('.doc-row').forEach(row => {
    if (row.querySelector('.row-checkbox:checked')) {
      row.classList.add('selected');
    } else {
      row.classList.remove('selected');
    }
  });
}

function toggleSelectAll(checkbox) {
  document.querySelectorAll('.row-checkbox').forEach(cb => {
    cb.checked = checkbox.checked;
  });
  updateSelection();
}

function clearSelection() {
  document.querySelectorAll('.row-checkbox').forEach(cb => {
    cb.checked = false;
  });
  document.getElementById('selectAllCheckbox').checked = false;
  updateSelection();
}

function confirmDelete() {
  const selectedIds = getSelectedDocIds();
  if (selectedIds.length === 0) {
    alert('No documents selected');
    return;
  }

  document.getElementById('deleteCount').textContent = selectedIds.length;
  document.getElementById('deleteConfirmModal').classList.add('active');
}

function cancelDelete() {
  document.getElementById('deleteConfirmModal').classList.remove('active');
}

function executeDelete() {
  const selectedIds = getSelectedDocIds();
  if (selectedIds.length === 0) return;

  // Send delete request to backend
  fetch('/delete-selected', {
    method: 'POST',
    headers: {
      'Content-Type': 'application/json'
    },
    body: JSON.stringify({
      documentIds: selectedIds
    })
  })
  .then(response => response.json())
  .then(data => {
    if (data.success) {
      alert(`✓ ${data.deleted} document(s) deleted successfully`);
      cancelDelete();
      clearSelection();
      location.reload(); // Reload to show updated list
    } else {
      alert(`Error: ${data.error || 'Failed to delete documents'}`);
    }
  })
  .catch(error => {
    console.error('Delete error:', error);
    alert('Error deleting documents');
  });
}

// Close modal when clicking outside
document.getElementById('deleteConfirmModal').addEventListener('click', function(e) {
  if (e.target === this) cancelDelete();
});

// ── PDF Modal Logic ──
function openModal(docId, s3Key, vendor, date, amount, conf, pdfUrl) {
  document.getElementById('modalTitle').textContent = s3Key;
  document.getElementById('mVendor').textContent    = vendor;
  document.getElementById('mDate').textContent      = date;
  document.getElementById('mAmount').textContent    = amount;
  document.getElementById('mConf').textContent      = conf;
  document.getElementById('approveId').value        = docId;
  document.getElementById('rejectId').value         = docId;

  // Load PDF in iframe if URL exists
  if (pdfUrl && pdfUrl !== 'None') {
    document.getElementById('pdfFrame').src = pdfUrl;
  } else {
    document.getElementById('pdfFrame').src =
      'data:text/html,<p style="font-family:Arial;padding:40px;color:#888">' +
      'PDF preview unavailable. Check S3 bucket permissions.</p>';
  }

  document.getElementById('pdfModal').classList.add('active');
}

function closeModal() {
  document.getElementById('pdfModal').classList.remove('active');
  document.getElementById('pdfFrame').src = '';
}

// Close modal when clicking outside
document.getElementById('pdfModal').addEventListener('click', function(e) {
  if (e.target === this) closeModal();
});

// ── Tab Switching Logic ──
function switchTab(tabName) {
  // Hide all tabs
  document.querySelectorAll('.tab-content').forEach(tab => {
    tab.classList.remove('active');
  });
  
  // Deactivate all buttons
  document.querySelectorAll('.tab-button').forEach(btn => {
    btn.classList.remove('active');
  });
  
  // Show selected tab
  document.getElementById(tabName).classList.add('active');
  
  // Activate corresponding button
  event.target.classList.add('active');
}

// ── File Upload Logic ──
const MAX_FILES = 3;
const MAX_FILE_SIZE = 10 * 1024 * 1024; // 10 MB
const selectedFiles = [];

function handleFileSelect(event) {
  const files = Array.from(event.target.files);
  
  // Clear previous selection
  selectedFiles.length = 0;
  
  // Validate and add files
  const errors = [];
  
  if (files.length > MAX_FILES) {
    errors.push(`Maximum ${MAX_FILES} files allowed. You selected ${files.length}.`);
  }
  
  for (const file of files) {
    if (!file.type.includes('pdf')) {
      errors.push(`${file.name} is not a PDF file.`);
      continue;
    }
    if (file.size > MAX_FILE_SIZE) {
      errors.push(`${file.name} exceeds 10 MB limit (${(file.size / 1024 / 1024).toFixed(2)} MB).`);
      continue;
    }
    if (selectedFiles.length < MAX_FILES) {
      selectedFiles.push(file);
    }
  }
  
  // Show errors if any
  if (errors.length > 0) {
    showUploadError(errors.join(' '));
  } else {
    hideUploadError();
  }
  
  // Render file list
  renderFileList();
  
  // Update button state
  document.getElementById('uploadBtn').disabled = selectedFiles.length === 0;
  document.getElementById('fileCountInfo').style.display = selectedFiles.length > 0 ? 'block' : 'none';
  document.getElementById('selectedFileCount').textContent = selectedFiles.length;
}

function renderFileList() {
  const fileListDiv = document.getElementById('fileList');
  fileListDiv.innerHTML = '';
  
  selectedFiles.forEach((file, index) => {
    const fileSize = (file.size / 1024 / 1024).toFixed(2);
    const fileItem = document.createElement('div');
    fileItem.className = 'file-item';
    fileItem.innerHTML = `
      <div class="file-item-info">
        <div class="file-item-name">📄 ${file.name}</div>
        <div class="file-item-size">${fileSize} MB</div>
      </div>
      <button type="button" class="file-item-remove" onclick="removeFile(${index})">
        Remove
      </button>
    `;
    fileListDiv.appendChild(fileItem);
  });
}

function removeFile(index) {
  selectedFiles.splice(index, 1);
  renderFileList();
  document.getElementById('uploadBtn').disabled = selectedFiles.length === 0;
  document.getElementById('fileCountInfo').style.display = selectedFiles.length > 0 ? 'block' : 'none';
  document.getElementById('selectedFileCount').textContent = selectedFiles.length;
  document.getElementById('pdfFileInput').value = '';
  hideUploadError();
}

function uploadFiles() {
  if (selectedFiles.length === 0) {
    showUploadError('No files selected');
    return;
  }
  
  // Show progress bar
  document.getElementById('uploadProgress').classList.add('active');
  hideUploadSuccess();
  hideUploadError();
  document.getElementById('uploadBtn').disabled = true;
  
  // Create FormData
  const formData = new FormData();
  selectedFiles.forEach(file => {
    formData.append('files', file);
  });
  
  // Upload with progress tracking
  const xhr = new XMLHttpRequest();
  
  // Progress event
  xhr.upload.addEventListener('progress', (e) => {
    if (e.lengthComputable) {
      const percentComplete = (e.loaded / e.total) * 100;
      document.getElementById('progressFill').style.width = percentComplete + '%';
      document.getElementById('progressText').textContent = Math.round(percentComplete) + '%';
    }
  });
  
  // Completion
  xhr.addEventListener('load', () => {
    document.getElementById('uploadProgress').classList.remove('active');
    
    if (xhr.status === 200) {
      const response = JSON.parse(xhr.responseText);
      if (response.success) {
        showUploadSuccess(`${response.uploaded} file(s) uploaded successfully!`);
        selectedFiles.length = 0;
        document.getElementById('fileList').innerHTML = '';
        document.getElementById('pdfFileInput').value = '';
        document.getElementById('fileCountInfo').style.display = 'none';
        document.getElementById('uploadBtn').disabled = true;
        document.getElementById('progressFill').style.width = '0%';
        document.getElementById('progressText').textContent = '0%';
      } else {
        showUploadError(response.error || 'Upload failed');
        document.getElementById('uploadBtn').disabled = false;
      }
    } else {
      showUploadError('Upload failed. Please try again.');
      document.getElementById('uploadBtn').disabled = false;
    }
  });
  
  // Error
  xhr.addEventListener('error', () => {
    document.getElementById('uploadProgress').classList.remove('active');
    showUploadError('Network error occurred');
    document.getElementById('uploadBtn').disabled = false;
  });
  
  xhr.open('POST', '/upload-pdf');
  xhr.send(formData);
}

function showUploadSuccess(message) {
  document.getElementById('successMessage').textContent = message;
  document.getElementById('uploadSuccess').classList.add('active');
}

function hideUploadSuccess() {
  document.getElementById('uploadSuccess').classList.remove('active');
}

function showUploadError(message) {
  document.getElementById('errorMessage').textContent = message;
  document.getElementById('uploadError').classList.add('active');
}

function hideUploadError() {
  document.getElementById('uploadError').classList.remove('active');
}
</script>

</body>
</html>
'''

def get_all_docs():
    result = table.scan()
    items  = result.get('Items', [])
    # Generate presigned URL for flagged docs
    for doc in items:
        if doc.get('flaggedForReview') and doc.get('status') == 'PENDING':
            doc['pdfUrl'] = get_presigned_url(doc.get('s3Key', ''))
        else:
            doc['pdfUrl'] = None
    return items

def sort_docs_by_date(docs, sort_order='desc'):
    """
    Sort documents by date in ascending or descending order.
    Handles various date formats including those with time components.
    """
    def parse_date(date_str):
        if not date_str or date_str == 'UNKNOWN':
            return datetime.min if sort_order == 'desc' else datetime.max
        
        try:
            # Try parsing common date formats
            for fmt in [
                '%d %B %Y %H:%M:%S',  # 15 January 2024 14:30:45
                '%d %B %Y %H:%M',      # 15 January 2024 14:30
                '%d %B %Y',            # 15 January 2024
                '%Y-%m-%d %H:%M:%S',   # 2024-01-15 14:30:45
                '%Y-%m-%d %H:%M',      # 2024-01-15 14:30
                '%Y-%m-%d',            # 2024-01-15
                '%m/%d/%Y %H:%M:%S',   # 01/15/2024 14:30:45
                '%m/%d/%Y',            # 01/15/2024
                '%d/%m/%Y %H:%M:%S',   # 15/01/2024 14:30:45
                '%d/%m/%Y',            # 15/01/2024
            ]:
                try:
                    return datetime.strptime(date_str.strip(), fmt)
                except ValueError:
                    continue
            # If no format matched, return a default value
            return datetime.min if sort_order == 'desc' else datetime.max
        except Exception:
            return datetime.min if sort_order == 'desc' else datetime.max
    
    sorted_docs = sorted(
        docs,
        key=lambda x: parse_date(x.get('date', '')),
        reverse=(sort_order == 'desc')
    )
    return sorted_docs

def calc_stats(docs):
    return {
        'total':    len(docs),
        'approved': sum(1 for d in docs if d.get('status') == 'APPROVED'),
        'pending':  sum(1 for d in docs if d.get('status') == 'PENDING'),
        'rejected': sum(1 for d in docs if d.get('status') == 'REJECTED'),
    }

@app.route('/')
def index():
    sort_order = request.args.get('sort', 'desc')  # Default to descending
    if sort_order not in ['asc', 'desc']:
        sort_order = 'desc'
    docs  = get_all_docs()
    docs  = sort_docs_by_date(docs, sort_order)
    stats = calc_stats(docs)
    return render_template_string(TEMPLATE, docs=docs, query=None, sort_order=sort_order, **stats)

@app.route('/search')
def search():
    q           = request.args.get('q', '').strip()
    sort_order  = request.args.get('sort', 'desc')  # Default to descending
    if sort_order not in ['asc', 'desc']:
        sort_order = 'desc'
    docs  = get_all_docs()
    if q:
        docs = [d for d in docs if q.lower() in str(d).lower()]
    docs  = sort_docs_by_date(docs, sort_order)
    stats = calc_stats(docs)
    return render_template_string(TEMPLATE, docs=docs, query=q, sort_order=sort_order, **stats)

@app.route('/approve', methods=['POST'])
def approve():
    doc_id = request.form['id']
    table.update_item(
        Key={'documentId': doc_id},
        UpdateExpression='SET #s = :v, flaggedForReview = :f',
        ExpressionAttributeNames={'#s': 'status'},
        ExpressionAttributeValues={':v': 'APPROVED', ':f': False}
    )
    return redirect('/')

@app.route('/reject', methods=['POST'])
def reject():
    doc_id = request.form['id']
    reason = request.form.get('reason', 'No reason provided').strip()
    table.update_item(
        Key={'documentId': doc_id},
        UpdateExpression='SET #s = :v, rejectReason = :r, flaggedForReview = :f',
        ExpressionAttributeNames={'#s': 'status'},
        ExpressionAttributeValues={
            ':v': 'REJECTED',
            ':r': reason,
            ':f': False
        }
    )
    return redirect('/')

@app.route('/delete-selected', methods=['POST'])
def delete_selected():
    """
    Delete multiple documents from DynamoDB and their PDFs from S3.
    Expects JSON body with 'documentIds' array.
    """
    try:
        data = request.get_json()
        document_ids = data.get('documentIds', [])

        if not document_ids:
            return jsonify({'success': False, 'error': 'No document IDs provided'}), 400

        deleted_count = 0
        errors = []

        # Delete each document
        for doc_id in document_ids:
            try:
                # First, get the document to retrieve the S3 key
                response = table.get_item(Key={'documentId': doc_id})
                doc = response.get('Item', {})
                s3_key = doc.get('s3Key')

                # Delete PDF from S3 if s3Key exists
                if s3_key:
                    try:
                        s3_client.delete_object(
                            Bucket=BUCKET_NAME,
                            Key=s3_key
                        )
                    except ClientError as e:
                        errors.append(f"Failed to delete S3 file for {doc_id}: {str(e)}")

                # Delete document from DynamoDB
                table.delete_item(
                    Key={'documentId': doc_id}
                )
                deleted_count += 1
            except ClientError as e:
                errors.append(f"Failed to delete {doc_id}: {str(e)}")

        response = {
            'success': True,
            'deleted': deleted_count,
            'total': len(document_ids)
        }

        if errors:
            response['errors'] = errors
            response['success'] = deleted_count == len(document_ids)

        return jsonify(response)

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/upload-pdf', methods=['POST'])
def upload_pdf():
    """
    Handle PDF file uploads to S3.
    Accepts up to 3 PDF files, max 10MB each.
    """
    try:
        files = request.files.getlist('files')

        if not files or len(files) == 0:
            return jsonify({'success': False, 'error': 'No files provided'}), 400

        # Validate file count
        if len(files) > 3:
            return jsonify({'success': False, 'error': 'Maximum 3 files allowed'}), 400

        uploaded_count = 0
        errors = []
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

        for file in files:
            if file.filename == '':
                continue

            # Validate file type
            if not file.filename.lower().endswith('.pdf'):
                errors.append(f"{file.filename} is not a PDF file")
                continue

            # Validate file size
            file.seek(0, 2)  # Seek to end
            file_size = file.tell()
            file.seek(0)  # Reset to beginning

            if file_size > MAX_FILE_SIZE:
                errors.append(f"{file.filename} exceeds 10 MB limit")
                continue

            try:
                # Generate secure filename with timestamp
                filename = secure_filename(file.filename)
                timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                unique_id = str(uuid.uuid4())[:8]
                s3_key = f"uploads/{timestamp}_{unique_id}_{filename}"

                # Upload to S3
                s3_client.put_object(
                    Bucket=BUCKET_NAME,
                    Key=s3_key,
                    Body=file.read(),
                    ContentType='application/pdf',
                    Metadata={
                        'original-filename': filename,
                        'upload-timestamp': datetime.utcnow().isoformat(),
                        'source': 'web-upload'
                    }
                )

                uploaded_count += 1

            except ClientError as e:
                errors.append(f"Failed to upload {file.filename}: {str(e)}")
            except Exception as e:
                errors.append(f"Error uploading {file.filename}: {str(e)}")

        response = {
            'success': uploaded_count > 0,
            'uploaded': uploaded_count,
            'total': len(files)
        }

        if errors:
            response['errors'] = errors

        return jsonify(response)

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/download-excel')
def download_excel():
    """
    Generate and download all documents as an Excel file.
    """
    try:
        # Get all documents
        docs = get_all_docs()
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Documents"
        
        # Define header style
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define data style
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Add headers
        headers = ["Doc ID", "File Name", "Vendor", "Date", "Amount", "Confidence", "Status", "Uploaded", "Reject Reason"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for row_num, doc in enumerate(docs, 2):
            row_data = [
                doc.get('documentId', '')[:8] + '...',
                doc.get('s3Key', ''),
                doc.get('vendor', ''),
                doc.get('date', ''),
                doc.get('amount', ''),
                str(doc.get('confidence', '')) + '%',
                doc.get('status', ''),
                doc.get('uploadedAt', '')[:10] if doc.get('uploadedAt') else '',
                doc.get('rejectReason', '')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = data_alignment
                cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 12  # Doc ID
        ws.column_dimensions['B'].width = 35  # File Name
        ws.column_dimensions['C'].width = 18  # Vendor
        ws.column_dimensions['D'].width = 15  # Date
        ws.column_dimensions['E'].width = 12  # Amount
        ws.column_dimensions['F'].width = 12  # Confidence
        ws.column_dimensions['G'].width = 12  # Status
        ws.column_dimensions['H'].width = 12  # Uploaded
        ws.column_dimensions['I'].width = 30  # Reject Reason
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        # Create a BytesIO object to hold the Excel file
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename with timestamp
        filename = f"JusticeArch_Documents_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)